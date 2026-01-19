import os
import sys
import json
import shutil
import time
from pathlib import Path
from pygltflib import GLTF2
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    openpyxl = None

# Supported file extensions
IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp'}
VIDEO_EXTENSIONS = {'.mp4', '.avi', '.mov', '.mkv', '.webm'}

# ============================================================================
# GLB HIERARCHY
# ============================================================================

# ============================================================================
# GLB HIERARCHY
# ============================================================================

def extract_hierarchy(glb_path):
    """Extract hierarchical structure from a GLB file."""
    try:
        gltf = GLTF2().load(glb_path)
        nodes_dict = {}
        if gltf.nodes:
            for idx, node in enumerate(gltf.nodes):
                node_name = node.name if node.name else f"Node_{idx}"
                nodes_dict[idx] = {'name': node_name, 'children': node.children or []}
        
        all_children = set()
        for n in nodes_dict.values(): all_children.update(n['children'])
        root_nodes = [i for i in nodes_dict if i not in all_children]
        
        def build(idx):
            if idx not in nodes_dict: return None
            node = nodes_dict[idx]
            children = [c for c in (build(child) for child in node['children']) if c]
            return {
                'name': node['name'],
                'children': children
            }
        
        hierarchy = []
        for root in root_nodes:
            # IMPORTANT: If the root node represents the model itself, we want its CHILDREN, not the root.
            # Heuristic: If there is only 1 root node, treat it as the "Container" and use its children.
            # If there are multiple roots, they are likely sibling parts.
            if len(root_nodes) == 1:
                root_tree = build(root)
                if root_tree and root_tree['children']:
                    hierarchy.extend(root_tree['children'])
                # If root has no children, return empty hierarchy (no parts)
            else:
                tree = build(root)
                if tree and tree['name'].strip(): hierarchy.append(tree)
                
        return hierarchy
    except Exception as e:
        print(f"❌ Error reading GLB: {e}")
        return []

def create_folders_from_hierarchy(hierarchy, parent_folder):
    """Recursively create ModelParts folders."""
    created = []
    for node in hierarchy:
        name = node['name']
        if not name.strip(): continue
        
        # Create folder for part
        part_folder = parent_folder / name
        part_folder.mkdir(exist_ok=True)
        created.append(part_folder)
        print(f"✔️  Created part: {part_folder}")
        
        # Subfolders
        (part_folder / "Button_Images").mkdir(exist_ok=True)
        (part_folder / "Video").mkdir(exist_ok=True)
        desc_folder = part_folder / "Description"
        desc_folder.mkdir(exist_ok=True)
        
        # Excel
        create_excel(desc_folder, f"DescriptionFor{name}")
        
        # Children
        if node['children']:
            created.extend(create_folders_from_hierarchy(node['children'], part_folder))
    return created

def create_excel(folder, name):
    if not openpyxl: return
    path = folder / f"{name}.xlsx"
    if path.exists(): return
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Description"
        ws['A1'], ws['B1'] = "Key", "Value"
        ws['A1'].font = ws['B1'].font = Font(bold=True)
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        wb.save(path)
        print(f"✔️  Created Excel: {path}")
    except Exception as e:
        print(f"❌ Error creating Excel: {e}")


# ============================================================================
# DATA GENERATION
# ============================================================================

def scan_images(folder, base):
    """Scan Button_Images folder."""
    buttons = []
    img_folder = folder / "Button_Images"
    # print(f"  Scanning images in: {img_folder}") # Debug
    if img_folder.exists():
        for f in img_folder.iterdir():
            if f.suffix.lower() in IMAGE_EXTENSIONS:
                try:
                    rel = f.relative_to(base)
                    buttons.append({
                        "name": f.stem,
                        "imagePath": str(rel).replace('\\', '/')
                    })
                except ValueError:
                    print(f"⚠️ Path error: {f} is not relative to {base}")
                    
    buttons.sort(key=lambda x: x['name'])
    return buttons

def scan_video(folder, base):
    """Scan Video folder."""
    vid_folder = folder / "Video"
    if vid_folder.exists():
        for f in vid_folder.iterdir():
            if f.suffix.lower() in VIDEO_EXTENSIONS:
                try:
                    return str(f.relative_to(base)).replace('\\', '/')
                except ValueError: pass
    return ""

def read_excel_file(path):
    if not openpyxl: return []
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2):
            k = str(row[0].value).strip() if row[0].value else ""
            v = str(row[1].value).strip() if row[1].value else ""
            if k or v: data.append({"key": k, "value": v})
        return data
    except Exception: return []

def read_desc(folder):
    """Read Excel from Description folder."""
    desc_folder = folder / "Description"
    if desc_folder.exists():
        for f in desc_folder.glob("*.xlsx"):
            if not f.name.startswith('~$'):
                return read_excel_file(f)
    return []

def scan_parts(folder, base):
    """Recursively scan ModelParts."""
    parts = []
    
    for item in folder.iterdir():
        if item.is_dir():
            if item.name in ["Button_Images", "Video", "Description"]: continue
            
            part = {
                "name": item.name,
                "video": scan_video(item, base),
                "datasheet": "",
                "description": read_desc(item),
                "buttons": scan_images(item, base),
                "parts": scan_parts(item, base)
            }
            parts.append(part)
    parts.sort(key=lambda x: x['name'])
    return parts

def generate_json(folder):
    path = Path(folder).resolve() # Ensure absolute path
    if not path.exists(): 
        print(f"❌ Path not found: {path}")
        return None
    
    print(f"🔄 Generating Data.json for: {path.name}...")
    
    infos = path / "ModelInfos"
    parts_dir = path / "ModelParts"
    
    # Check if critical folders match expectations
    if not infos.exists():
        print(f"⚠️  ModelInfos not found at {infos}")
    
    # Model File
    model_url = ""
    glb_dir = infos / "3DMODEL"
    if glb_dir.exists():
        glbs = list(glb_dir.glob("*.glb"))
        if glbs: 
            try:
                model_url = str(glbs[0].relative_to(path)).replace('\\', '/')
            except ValueError: pass
    if not model_url:
        # Fallback
        glbs = list(path.glob("*.glb"))
        if glbs: model_url = glbs[0].name
        
    # QR Code
    qr_url = "QRCode.png"
    qr_dir = infos / "QRCode"
    if qr_dir.exists():
        imgs = [f for f in qr_dir.iterdir() if f.suffix.lower() in IMAGE_EXTENSIONS]
        if imgs: 
            try:
                qr_url = str(imgs[0].relative_to(path)).replace('\\', '/')
            except ValueError: pass
        
    data = [{
        "name": path.name,
        "qr_image_url": qr_url,
        "modelFileUrl": model_url,
        "video": scan_video(infos, path),
        "Description": read_desc(infos),
        "Buttons": scan_images(infos, path),
        "parts": scan_parts(parts_dir, path) if parts_dir.exists() else []
    }]
    
    json_path = path / "Data.json"
    try:
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        print(f"✅ Saved Data.json (Size: {json_path.stat().st_size} bytes)")
    except Exception as e:
        print(f"❌ Error writing Data.json: {e}")

class Handler(FileSystemEventHandler):
    def on_created(self, event):
        if event.is_directory: return
        self._process(Path(event.src_path))
        
    def on_modified(self, event):
        if event.is_directory: return
        self._process(Path(event.src_path))

    def _process(self, path):
        path = path.resolve()
        
        # 1. New GLB detection
        if path.suffix.lower() == '.glb':
            is_in_structure = "3DMODEL" in [p.name for p in path.parents]
            if not is_in_structure:
                print(f"🆕 New GLB detected: {path.name}")
                time.sleep(1) 
                create_structure(path)
                return

        # 2. Batch file auto-moving
        if path.suffix.lower() == '.bat':
            # If a .bat is dropped at root, move it to the first available project folder
            root_dir = Path(".").resolve()
            if path.parent == root_dir:
                # Find project folders
                for p in root_dir.iterdir():
                    if p.is_dir() and ((p / "ModelInfos").exists() or (p / "Data.json").exists()):
                        print(f"📦 Moving script {path.name} to project folder: {p.name}")
                        try:
                            # Wait a bit for file to be ready
                            time.sleep(0.5)
                            dest = p / path.name
                            shutil.move(str(path), str(dest))
                            return # Only move to the first one found
                        except Exception as e:
                            print(f"❌ Error moving .bat: {e}")

        # 3. Existing project updates
        if path.suffix.lower() in IMAGE_EXTENSIONS | VIDEO_EXTENSIONS | {'.xlsx', '.glb'}:
            p = path.parent
            for _ in range(5):
                if len(p.parts) < 2: break
                if (p / "ModelInfos").exists() or (p / "Data.json").exists():
                    if path.name == "Data.json": return 
                    print(f"📝 Change in {path.name} -> Updating directory: {p.name}")
                    try:
                        time.sleep(1) 
                        generate_json(p)
                    except Exception as e:
                        print(f"❌ Error regenerating: {e}")
                    break
                p = p.parent

def create_structure(glb_path):
    glb = Path(glb_path)
    if not glb.exists() or glb.suffix.lower() != '.glb': return
    
    root = glb.parent / glb.stem
    # If it's already a folder, we might be re-triggering on a move.
    # But watcher check 'is_in_structure' should prevent this.
    
    print(f"📂 Creating structure for {root.name}...")
    root.mkdir(exist_ok=True)
    
    # ModelInfos
    infos = root / "ModelInfos"
    infos.mkdir(exist_ok=True)
    (infos / "3DMODEL").mkdir(exist_ok=True)
    (infos / "Button_Images").mkdir(exist_ok=True)
    (infos / "Video").mkdir(exist_ok=True)
    (infos / "Description").mkdir(exist_ok=True)
    (infos / "QRCode").mkdir(exist_ok=True)
    
    create_excel(infos / "Description", f"DescriptionFor{root.name}")
    
    # Move GLB
    dest_glb = infos / "3DMODEL" / glb.name
    try:
        shutil.move(str(glb), str(dest_glb))
    except Exception as e:
        print(f"❌ Error moving GLB to final location: {e}")
    
    # ModelParts
    hier = extract_hierarchy(str(dest_glb))
    if hier:
        parts = root / "ModelParts"
        parts.mkdir(exist_ok=True)
        create_folders_from_hierarchy(hier, parts)
        
    generate_json(root)

def watch(folder):
    obs = Observer()
    obs.schedule(Handler(), str(folder), recursive=True)
    obs.start()
    print(f"👀 Watching {folder} for changes...")
    try:
        while True: time.sleep(1)
    except KeyboardInterrupt:
        obs.stop()
    obs.join()

if __name__ == "__main__":
    if len(sys.argv) > 1:
        arg = Path(sys.argv[1])
        if arg.suffix.lower() == '.glb': create_structure(arg)
        elif arg.is_dir(): generate_json(arg)
    else:
        # Watch current dir
        watch(Path("."))
